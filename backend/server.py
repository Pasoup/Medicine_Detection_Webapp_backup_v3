# =============================================================================
#  backend/server.py — FastAPI bridge between the web frontend and pipeline
#
#  Run with:  uvicorn backend.server:app --host 0.0.0.0 --port 8000 --reload
# =============================================================================

import sys
import os
import json
import base64
import tempfile
import uuid
from datetime import datetime, timedelta
from typing import Optional
import time
import threading

import cv2
from concurrent.futures import ThreadPoolExecutor
import numpy as np
from fastapi import FastAPI, HTTPException, Request, Depends
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.exceptions import RequestValidationError
from fastapi.responses import JSONResponse, StreamingResponse
from fastapi.security import OAuth2PasswordBearer

from pydantic import BaseModel

import camera as cam_module
from database import init_db, get_db,pwd_context
from jose import jwt, JWTError

# ── Add project root so pipeline imports work ─────────────────────────────────
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from config import BOX_CLASS_ID, QR_CLASS_ID, LOG_DIR, L1_PAD_QR
from utils.medicine_db import load_medicine_db, db_match, normalize_ocr, db_confidence_tier
from utils.image import get_perspective
from pipeline.layer1_detect import load_layer1_model, layer1_detect
from pipeline.layer2_qr import layer2_read_qr
from pipeline.layer3_ocr import layer3_read_label
from pipeline.layer4_vision import layer4_scan_full_frame, layer4_match_to_box, LAYER4_CLASS_NAMES
from pipeline.consensus import consensus_check
from led import led_green, led_off, led_orange, led_red, led_white, led_unknown, connect as led_connect, is_connected as led_is_connected


os.makedirs(LOG_DIR, exist_ok=True)
init_db()
led_connect()

SECRET_KEY = 'abefkjhaekjfhkajef'
ALGORITHM = "HS256"
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/auth/login")
# ── Calibration (written by image_taking/calibration.py) ──────────────────────
_CALIB_PATH = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    "..", "frontend", "src","utils", "calibration.json"
)
_CALIB_DEFAULTS = {"crop0_right": 0, "crop1_left": 0, "y_offset": 0, "x_offset": 0}

def load_calibration() -> dict:
    """Read calibration.json from disk each time — picks up edits without restart."""
    if os.path.exists(_CALIB_PATH):
        try:
            with open(_CALIB_PATH) as _f:
                return {**_CALIB_DEFAULTS, **json.load(_f)}
        except Exception as e:
            print(f"  Calibration: read error — {e}")
    return _CALIB_DEFAULTS.copy()

print(f"  Calibration path : {os.path.abspath(_CALIB_PATH)}")
print(f"  Calibration file exists: {os.path.exists(_CALIB_PATH)}")
print(f"  Calibration: {load_calibration()}")

def blink_then_white(color_fn, blinks=3, delay=1.5):
    for _ in range(blinks):
        time.sleep(0.3)
        led_off()
        time.sleep(0.3)
        color_fn()
    time.sleep(delay)
    led_white()

# ── App setup ─────────────────────────────────────────────────────────────────
app = FastAPI(title="MedVerify API", version="2.0")

@app.exception_handler(RequestValidationError)
async def validation_error_handler(request: Request, exc: RequestValidationError):
    print("\n" + "═" * 50)
    print("  422 — REQUEST BODY REJECTED")
    for err in exc.errors():
        print(f"  field : {err.get('loc')}")
        print(f"  error : {err.get('msg')}")
        print(f"  type  : {err.get('type')}")
    try:
        body = await request.body()
        preview = body[:300].decode("utf-8", errors="replace")
        print(f"  body  : {preview}…")
    except Exception:
        pass
    print("═" * 50 + "\n")
    return JSONResponse(status_code=422, content={"detail": exc.errors()})

app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:3000", "http://localhost:5173"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ── Serve medicine reference images as static files ───────────────────────────
MEDICINE_IMAGES_DIR = os.path.join(
    os.path.dirname(os.path.abspath(__file__)), "data", "medicine_images"
)
os.makedirs(MEDICINE_IMAGES_DIR, exist_ok=True)
app.mount("/medicine-images", StaticFiles(directory=MEDICINE_IMAGES_DIR),
          name="medicine-images")


def find_medicine_image(name: str) -> str | None:
    
    if not name or not os.path.exists(MEDICINE_IMAGES_DIR):
        return None

    def norm(s: str) -> str:
        return s.lower().replace(" ", "").replace("-", "").replace("_", "")

    name_norm = norm(name)
    extensions = {".jpg", ".jpeg", ".png", ".webp", ".bmp"}

    for fname in os.listdir(MEDICINE_IMAGES_DIR):
        stem, ext = os.path.splitext(fname)
        if ext.lower() not in extensions:
            continue
        if norm(stem) == name_norm:
            return f"/medicine-images/{fname}"

    return None

# ── Load models once at startup ───────────────────────────────────────────────
print("═" * 50)
print("  MEDVERIFY API — INITIALISING")
print("═" * 50)

layer1_model = load_layer1_model()
medicine_db  = load_medicine_db(yolo_class_names=LAYER4_CLASS_NAMES)

print(f"  Medicine DB: {len(medicine_db)} entries loaded.")
_cam_ok = cam_module.start(cam0_index=0, cam1_index=1)
if not _cam_ok:
    print("  [WARNING] Camera failed to start — running without camera.")
print("  Ready.")

# ── In-memory state (replace with a real DB in production) ───────────────────
expected_list: list[dict] = []   # [{"name": str, "quantity": int}]
scan_history:  list[dict] = []   # list of past scan result summaries


# ── Request / Response models ─────────────────────────────────────────────────

class ScanRequest(BaseModel):
    frame_b64:  str  = ""    # single stitched frame (correct)
    frames_b64: list = []    # array of frames (fallback if frontend sends this)
    expected:   list = []

class MedicineItem(BaseModel):
    name:     str
    quantity: int = 1


class ExpectedListPayload(BaseModel):
    medicines: list[MedicineItem]

class CalibrationUpdate(BaseModel):
    cam_width:   int | None = None
    cam_height:  int | None = None
    crop0_right: int | None = None
    crop1_left:  int | None = None
    y_offset:    int | None = None
    x_offset:    int | None = None

class LoginRequest(BaseModel):
    username: str
    password: str


# ---- Database ------------    
class DrugCreate(BaseModel):
    name: str

class DrugUpdate(BaseModel):
    name: str

class ScanResultItem(BaseModel):
    box_id:          str | None = None
    final_name:      str | None = None
    scan_status:     str | None = None
    confidence:      str | None = None
    ocr_raw:         str | None = None
    qr_name:         str | None = None
    reference_image: str | None = None

class SaveHistroyPayload(BaseModel):
    timestamp:      str 
    scanned_by:     str | None =None
    location:       str | None = None
    matched:        int = 0
    missing:        int = 0
    extra:          int = 0
    review:         int = 0
    unknown:        int = 0
    annotated:      str | None = None
    results:        list[ScanResultItem] = []
# ---- Med Codes------------

class MedCodes(BaseModel):
    label:      str    

class MedCodesItem(BaseModel):
    drug_name: str
    quantity: int = 1

class MedCodesItemUpdate(BaseModel):
    quantity: int

def _reload_medicine_db():
    global medicine_db
    with get_db() as conn:
        rows = conn.execute("SELECT name FROM drugs").fetchall()
    medicine_db = sorted(r["name"].upper() for r in rows)
    print(f"[DB] medicine_db reloaded - {len(medicine_db)} entries")


def get_current_user(token = Depends(oauth2_scheme)):
    try:
        payload = jwt.decode(token,SECRET_KEY,algorithms=[ALGORITHM])
        username = payload.get("sub")
        if username:
            return {"sub" : payload["sub"], "role" : payload["role"]}
        raise HTTPException(status_code=401, detail="no user found")
    except JWTError:
        raise HTTPException(status_code=401, detail="decode failed")

# ── Database Endpoint ─────────────────────────────────────────────────────────

@app.get("/drug-database")
def get_drug_database(current_user = Depends(get_current_user)):
    with get_db() as conn:
        rows = conn.execute("SELECT id, name FROM drugs ORDER BY name").fetchall()
    drugs = [{"id": r["id"], "name": r["name"]} for r in rows]
    return {"drugs": drugs, "total": len(drugs)}

@app.post("/drug-database")
def add_drug(payload: DrugCreate, current_user = Depends(get_current_user)):
    name = payload.name.strip()
    if not name:
        raise HTTPException(status_code=400, detail="Drug name cannot be empty")

    try:
        with get_db() as conn:
            conn.execute("INSERT INTO drugs (name) VALUES (?)", (name,))

    except Exception:
        raise HTTPException(status_code=400, detail=f"'{name}' already exists")
    _reload_medicine_db()
    return get_drug_database()

@app.delete("/drug-database/{drug_id}")
def delete_drug(drug_id: int, current_user = Depends(get_current_user)):
    with get_db() as conn:
        cur = conn.execute("DELETE FROM drugs WHERE id = ?", (drug_id,))
        if cur.rowcount == 0:
            raise HTTPException(status_code=404, detail=f"Drug ID: {drug_id} not found")
    _reload_medicine_db()
    return get_drug_database()

@app.put("/drug-database/{drug_id}")
def update_drug(drug_id: int, payload: DrugUpdate, current_user = Depends(get_current_user)):
    name = payload.name.strip()
    if not name:
        raise HTTPException(status_code=400, detail="Drug id cannot be empty")
    try:
        with get_db() as conn:
            cur = conn.execute("UPDATE drugs SET name = ? WHERE id = ?", (name,drug_id))
            
            if cur.rowcount == 0:
                raise HTTPException(status_code=404, detail=f"Drug ID {drug_id} not found")
    except HTTPException:
        raise
    except Exception:
        raise HTTPException(status_code=400, detail=f"'{name}' already exists")
    _reload_medicine_db()
    return get_drug_database()


# ── LED Toggle Endpoint ────────────────────────────────────────────────────────────
class LedToggle(BaseModel):
    on: bool

@app.post("/led/toggle")
def led_toggle(body: LedToggle, current_user = Depends(get_current_user)):
    if body.on:
        led_white()
    else:
        led_off()
    return {"on": body.on}


# ── MedicineCodes Endpoint ─────────────────────────────────────────────────────────

@app.get("/medicine-codes")
def get_medicine_codes(current_user = Depends(get_current_user)):
    with get_db() as conn:
        codes = conn.execute("SELECT * FROM medicine_codes").fetchall()
        result = []
        for code in codes:
            items = conn.execute(
                "SELECT id, drug_name, quantity FROM medicine_code_items WHERE code_id = ?",
                (code["id"],)
            ).fetchall()
            result.append({
                "id": code["id"],
                "label": code["label"],
                "items": [{"id": r["id"], "drug_name": r["drug_name"], "quantity": r["quantity"]} for r in items]
            })
    return result


@app.post("/medicine-codes")
def add_medicine_codes(payload : MedCodes,current_user = Depends(get_current_user)):
    name = payload.label.strip()

    if not name:
        raise HTTPException(status_code=400, detail="Label Name cannot be empty")

    with get_db() as conn:
        cur  = conn.execute("INSERT INTO medicine_codes(label) VALUES(?)", (name,))

    return( {"id" : cur.lastrowid, "label" : name})


@app.get("/medicine-codes/{med_code_id}")
def get_med_code_one(med_code_id: int,current_user = Depends(get_current_user)):
    with get_db() as conn:
        cur = conn.execute("SELECT * FROM medicine_codes WHERE id = ?", (med_code_id,)).fetchone()
        if not cur:
            raise HTTPException(status_code=404, detail="Medicine code not found")
        cur2 = conn.execute("SELECT * FROM medicine_code_items WHERE code_id = ?",(med_code_id,)).fetchall()

    return {
        "id" : cur["id"],
        "label" : cur["label"],
        "items" : [{"id" : row["id"], "drug_name" : row["drug_name"], "quantity" : row["quantity"]} for row in cur2]

    }

@app.put("/medicine-codes/{med_code_id}")
def update_medicine_code(med_code_id: int,payload: MedCodes, current_user = Depends(get_current_user)):
    name = payload.label.strip()
    if not name:
        raise HTTPException(status_code=400, detail="label name cannot be empty")
    
    with get_db() as conn:
        exist = conn.execute("SELECT id FROM medicine_codes WHERE id = ?", (med_code_id,)).fetchone()
        if not exist:
            raise HTTPException(status_code=404, detail="Med code doesn't exist")
        cur = conn.execute("UPDATE medicine_codes SET label = ? WHERE id = ?", (name,med_code_id))

    return {"id" : med_code_id, "label" : name}


@app.delete("/medicine-codes/{med_code_id}")
def delete_medicine_code(med_code_id: int, current_user = Depends(get_current_user)):
    with get_db() as conn:
        exist = conn.execute("SELECT id FROM medicine_codes WHERE id = ?", (med_code_id,)).fetchone()
        if not exist:
            raise HTTPException(status_code=404, detail="Med code not found")
        cur = conn.execute("DELETE FROM medicine_codes WHERE id = ?", (med_code_id,))
    return {"detail" : "deleted"}


# ── MedCode ITEMS Endpoint ─────────────────────────────────────────────────────────



@app.post("/medicine-codes/{med_code_id}/items")
def add_medcode_item(med_code_id : int, payload: MedCodesItem,current_user = Depends(get_current_user)):
    with get_db() as conn:
        exist = conn.execute("SELECT id FROM medicine_codes WHERE id = ?", (med_code_id,)).fetchone()
        if not exist:
            raise HTTPException(status_code=404, detail="Med code doesn't exist")
        existing_item = conn.execute(
            "SELECT id, quantity FROM medicine_code_items WHERE code_id = ? AND drug_name = ?",
            (med_code_id, payload.drug_name)
        ).fetchone()
        if existing_item:
            new_qty = existing_item["quantity"] + payload.quantity
            conn.execute("UPDATE medicine_code_items SET quantity = ? WHERE id = ?", (new_qty, existing_item["id"]))
            return {"id": existing_item["id"], "drug_name": payload.drug_name, "quantity": new_qty}
        cur = conn.execute("INSERT INTO medicine_code_items(code_id, drug_name, quantity) VALUES(?,?,?)", (med_code_id, payload.drug_name, payload.quantity))

    return {"id": cur.lastrowid, "drug_name": payload.drug_name, "quantity": payload.quantity}


@app.put("/medicine-codes/{med_code_id}/items/{item_id}")
def update_med_code_item(med_code_id: int, item_id: int, payload: MedCodesItemUpdate, current_user = Depends(get_current_user)):
    if payload.quantity < 1:
        raise HTTPException(status_code=400, detail="Quantity must be at least 1")
    with get_db() as conn:
        exist_item = conn.execute(
            "SELECT id FROM medicine_code_items WHERE id = ? AND code_id = ?", (item_id, med_code_id)
        ).fetchone()
        if not exist_item:
            raise HTTPException(status_code=404, detail="Item not found")
        conn.execute("UPDATE medicine_code_items SET quantity = ? WHERE id = ?", (payload.quantity, item_id))
    return {"id": item_id, "quantity": payload.quantity}


@app.delete("/medicine-codes/{med_code_id}/items/{item_id}")
def delete_med_code_item(med_code_id: int, item_id: int, current_user = Depends(get_current_user)):
    with get_db() as conn:
        exist_med_code = conn.execute("SELECT id FROM medicine_codes WHERE id = ?", (med_code_id, )).fetchone()
        if not exist_med_code:
            raise HTTPException(status_code=404, detail="Med Code doesn't exist")
        exist_item = conn.execute("SELECT id from medicine_code_items WHERE id = ?", (item_id,)).fetchone()
        if not exist_item:
            raise HTTPException(status_code=404, detail="item code doesn't exist")
        conn.execute("DELETE FROM medicine_code_items WHERE id = ?", (item_id,))
    return ({"detail" : "deleted"})




# ── Histroy Endpoint ─────────────────────────────────────────────────────────



@app.post("/history")
def save_history(payload: SaveHistroyPayload, current_user = Depends(get_current_user)):
    with get_db() as conn:
        cur = conn.execute("""
            INSERT INTO scan_sessions               
                (timestamp, scanned_by, location ,matched, missing, extra, review, unknown, 
annotated)
            VALUES (?,?,?,?,?,?,?,?,?)
            """, (payload.timestamp,
                  payload.scanned_by,
                  payload.location,
                  payload.matched,
                  payload.missing,
                  payload.extra,
                  payload.review,
                  payload.unknown,
                  payload.annotated,
        ))
        session_id = cur.lastrowid
        for r in payload.results:
            conn.execute(""" 
                INSERT INTO scan_results
                         (session_id, box_id, final_name, scan_status, confidence, ocr_raw,
                          qr_name, reference_image)
                VALUES (?,?,?,?,?,?,?,?)
            """, (
                session_id,
                r.box_id,
                r.final_name,
                r.scan_status,
                r.confidence,
                r.ocr_raw,
                r.qr_name,
                r.reference_image,
            ))
    return {"id" : session_id, "status" : "saved"}


@app.get("/history")
def get_history(current_user = Depends(get_current_user)):
    with get_db() as conn:
        session = conn.execute(""" 
            SELECT * FROM scan_sessions ORDER BY id DESC

""").fetchall()
    result = []

    for s in session:
        result.append({
            "id":   s["id"],
            "timestamp":    s["timestamp"],
            "scanned_by":   s["scanned_by"],
            "location":     s["location"],
            "matched":      s["matched"],
            "missing":      s["missing"],
            "extra":        s["extra"],
            "review":       s["review"],
            "unknown":      s["unknown"],
        })
    return {"history": result, "total": len(result)}


@app.get("/history/export")
def export_history(current_user = Depends(get_current_user)):
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from fastapi.responses import StreamingResponse

    with get_db() as conn:
        sessions = conn.execute("SELECT * FROM scan_sessions ORDER BY id DESC").fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Scan History"

    # Header styling
    header_font  = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    header_fill  = PatternFill("solid", fgColor="2563EB")
    center       = Alignment(horizontal="center", vertical="center")
    thin         = Side(style="thin", color="D1D5DB")
    border       = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["#", "Timestamp", "Scanned By", "Location",
               "Matched", "Missing", "Extra", "Unknown", "Status"]
    col_widths = [5, 22, 18, 20, 10, 10, 10, 10, 12]

    for col_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center
        cell.border    = border
        ws.column_dimensions[cell.column_letter].width = w

    ws.row_dimensions[1].height = 20

    # Status helpers
    def status_label(s):
        if s["missing"] and s["missing"] > 0: return "Failed"
        if (s["extra"] and s["extra"] > 0) or (s["unknown"] and s["unknown"] > 0): return "Partial"
        return "Passed"

    status_fill = {
        "Passed":  PatternFill("solid", fgColor="D1FAE5"),
        "Partial": PatternFill("solid", fgColor="FEF3C7"),
        "Failed":  PatternFill("solid", fgColor="FEE2E2"),
    }

    for row_idx, s in enumerate(sessions, start=2):
        label = status_label(s)
        row_data = [
            row_idx - 1,
            s["timestamp"],
            s["scanned_by"] or "",
            s["location"] or "",
            s["matched"]  or 0,
            s["missing"]  or 0,
            s["extra"]    or 0,
            s["unknown"]  or 0,
            label,
        ]
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border    = border
            cell.alignment = center
            if col_idx == 9:
                cell.fill = status_fill.get(label, PatternFill())

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"scan_history_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@app.get("/history/{session_id}")
def get_history_detail(session_id: int, current_user = Depends(get_current_user)):
    with get_db() as conn:
        session = conn.execute(
            "SELECT * FROM scan_sessions WHERE id = ?", (session_id,)
        ).fetchone()

        if not session:
            raise HTTPException(status_code=404, detail=f"Session {session_id} not found")
    
        results = conn.execute(
            "SELECT * FROM scan_results WHERE session_id = ?", (session_id,)
        ).fetchall()
    return {
        "id":         session["id"],
        "timestamp":  session["timestamp"],
        "scanned_by": session["scanned_by"],
        "matched":    session["matched"],
        "missing":    session["missing"],
        "extra":      session["extra"],
        "review":     session["review"],
        "unknown":    session["unknown"],
        "annotated":  session["annotated"],
        "results": [
            {
                "box_id":          r["box_id"],
                "final_name":      r["final_name"],
                "scan_status":     r["scan_status"],
                "confidence":      r["confidence"],
                "ocr_raw":         r["ocr_raw"],
                "qr_name":         r["qr_name"],
                "reference_image": r["reference_image"],
            }
            for r in results
        ]
    }

# ── Status Endpoint ───────────────────────────────────────────────────────────
@app.get("/status")
def get_status():
    return {
        "led": led_is_connected(),
        "camera": cam_module.is_running(),
    }

@app.post("/camera/restart")
def restart_camera(current_user = Depends(get_current_user)):
    cam_module.stop()
    ok = cam_module.start(cam0_index=0, cam1_index=1)
    if not ok:
        raise HTTPException(status_code=500, detail="Camera failed to restart")
    return {"ok": True}


#----LoginEndpoints ───────────────────────────────────────────────────────────────────
@app.post("/auth/login")
def loginRequest(request: LoginRequest):
    with get_db() as conn:
        res = conn.execute("SELECT * FROM users WHERE username = ?", (request.username,)).fetchone()
        if not res:
            raise HTTPException(status_code=404, detail="no user found")
        pas = pwd_context.verify(request.password, res["password_hash"])
        if not pas:
            raise HTTPException(status_code=401, detail="password worng")
        payload = {
            "sub" : res["username"],
            "role" : res["role"],
            "exp"  : datetime.utcnow() + timedelta(minutes=30)
        }
        token = jwt.encode(payload,SECRET_KEY,algorithm=ALGORITHM)
    

    return {"access_token": token, "token_type" : "bearer"}



#----UsersEndpoints ───────────────────────────────────────────────────────────────────

@app.get("/users")
def get_users(current_users = Depends(get_current_user)):
    if current_users["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")

    with get_db() as conn:
        rows = conn.execute("SELECT id, username, role FROM users").fetchall()
    return {"users" : [dict(r) for r in rows]}
    

@app.post("/users")
def create_user(request: dict, current_user = Depends(get_current_user)):
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")

    hashed = pwd_context.hash(request["password"])
    try:
        with get_db() as conn:
            conn.execute("INSERT INTO users(username,password_hash,role) VALUES(?,?,?)",
                (request["username"],hashed,request.get("role","pharmacist")))
            conn.commit()

    except Exception:
        raise HTTPException(status_code=400, detail="Username already exist")
    return {"ok": True}


@app.put("/users/{user_id}")
def update_users(user_id: int, request: dict, current_user = Depends(get_current_user)):
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access requried")
    
    with get_db() as conn:
        if "password" in request:
            hashed = pwd_context.hash(request["password"])
            conn.execute("UPDATE users SET username=?, role=?, password_hash=? WHERE id=?",
                         (request["username"],request["role"],hashed,user_id))
            
        else:
            conn.execute("UPDATE users SET username=?, role=? WHERE id=?", 
                         (request["username"], request["role"], user_id))
        conn.commit()
    return{"ok": True}


@app.delete("/users/{user_id}")
def delete_user(user_id: int, current_user = Depends(get_current_user)):
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    with get_db() as conn:
        conn.execute("DELETE FROM users WHERE id=?", (user_id,))
        conn.commit()
    return {"ok" : True}


# ── Helpers ───────────────────────────────────────────────────────────────────

def b64_to_frame(b64: str) -> np.ndarray:
    """Decode a base64 JPEG string to a BGR numpy array."""
    if "," in b64:
        b64 = b64.split(",", 1)[1]
    img_bytes = base64.b64decode(b64)
    arr = np.frombuffer(img_bytes, dtype=np.uint8)
    frame = cv2.imdecode(arr, cv2.IMREAD_COLOR)
    if frame is None:
        raise ValueError("Could not decode image from base64")
    return frame


@app.get("/calibration")
def get_calibration(current_user = Depends(get_current_user)):
    """Return the current camera calibration values to the frontend."""
    return load_calibration()


@app.post("/calibration")
def update_calibration(update: CalibrationUpdate, current_user = Depends(get_current_user)):
    """
    Write updated calibration values to calibration.json, then restart the
    cameras so the new resolution takes effect immediately.
    Only fields that are explicitly provided are updated; the rest are kept.
    """
    current = load_calibration()

    if update.cam_width   is not None: current["cam_width"]   = update.cam_width
    if update.cam_height  is not None: current["cam_height"]  = update.cam_height
    if update.crop0_right is not None: current["crop0_right"] = update.crop0_right
    if update.crop1_left  is not None: current["crop1_left"]  = update.crop1_left
    if update.y_offset    is not None: current["y_offset"]    = update.y_offset
    if update.x_offset    is not None: current["x_offset"]    = update.x_offset

    try:
        with open(_CALIB_PATH, "w") as f:
            json.dump(current, f, indent=2)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Could not write calibration.json: {e}")

    # Restart cameras so the new resolution is applied
    cam_module.stop()
    time.sleep(1.0)          # let DirectShow fully release the handles
    cam_module.start(cam0_index=0, cam1_index=1)

    return {"calibration": current, "restarted": True}


@app.get("/video_feed")
def video_feed():
    """MJPEG stream of the live stitched camera feed."""
    def generate():
        while True:
            frame = cam_module.get_jpeg()
            if frame:
                yield (
                    b"--frame\r\n"
                    b"Content-Type: image/jpeg\r\n\r\n"
                    + frame +
                    b"\r\n"
                )
            time.sleep(1 / 30)
    return StreamingResponse(
        generate(),
        media_type="multipart/x-mixed-replace; boundary=frame",
    )


def stitch_frames(frames: list, calib: dict | None = None) -> tuple:
    """
    Rotate each camera frame to match the display orientation, apply
    calibration offsets, then stitch side-by-side into a single wide image.

    Camera mounting orientation (matches CSS in CameraSection.jsx):
      cam0 → rotate -90° (CCW)
      cam1 → rotate +90° (CW)

    Returns (stitched, f0, f1).
    """
    c = calib or load_calibration()
    c0r  = c.get("crop0_right", 0)
    c1l  = c.get("crop1_left",  0)
    yOff = c.get("y_offset",    0)
    xOff = c.get("x_offset",    0)

    if len(frames) == 1:
        f0 = frames[0]
        return f0, f0, None

    f0 = cv2.rotate(frames[0], cv2.ROTATE_90_COUNTERCLOCKWISE)
    f1 = cv2.rotate(frames[1], cv2.ROTATE_90_CLOCKWISE)

    # Seam crop
    if c0r > 0: f0 = f0[:, :f0.shape[1] - c0r]
    if c1l > 0: f1 = f1[:, c1l:]

    # Vertical alignment
    if yOff > 0:
        f1 = f1[yOff:, :]
        f0 = f0[:f0.shape[0] - yOff, :]
    elif yOff < 0:
        y  = -yOff
        f0 = f0[y:, :]
        f1 = f1[:f1.shape[0] - y, :]

    # Horizontal alignment
    if xOff > 0:
        pad = np.zeros((f1.shape[0], xOff, 3), dtype=np.uint8)
        f1  = np.hstack([pad, f1])
    elif xOff < 0:
        f1  = f1[:, -xOff:]

    # Match heights (trim to shorter)
    h = min(f0.shape[0], f1.shape[0])
    f0, f1 = f0[:h], f1[:h]

    return cv2.hconcat([f0, f1]), f0, f1


def best_ocr_label(ocr_texts: list, threshold: float = 85.0) -> str:
    if not ocr_texts:
        return ""
    if len(ocr_texts) == 1:
        return ocr_texts[0]
    best_text, best_score = ocr_texts[0], 0.0
    for text in ocr_texts:
        if not text or not text.strip():
            continue
        _, score = db_match(normalize_ocr(text.lower()), medicine_db, threshold=threshold)
        if score > best_score:
            best_score, best_text = score, text
    return best_text


def frame_to_b64(frame: np.ndarray) -> str:
    """Encode a BGR numpy frame to base64 JPEG for sending back to the browser."""
    _, buf = cv2.imencode(".jpg", frame, [cv2.IMWRITE_JPEG_QUALITY, 80])
    return "data:image/jpeg;base64," + base64.b64encode(buf).decode()


def _get_qr_crop(det: dict,
                 f0: "np.ndarray | None",
                 f1: "np.ndarray | None") -> np.ndarray:
    """
    Re-extract a perspective-corrected QR crop from the individual camera frame.

    Layer 1 already supplies a crop from the stitched frame (det["crop"]).
    For QR reading we can do better: re-warp the *same* OBB points against the
    single-camera frame that the QR actually came from, which is never scaled or
    concatenated and therefore has no cross-camera seam artefacts.

    Logic
    -----
    The stitched frame is [f0 | f1] side by side.
    - bbox centre_x < f0.shape[1]  → QR is in cam0 (f0); pts are unchanged
    - bbox centre_x ≥ f0.shape[1]  → QR is in cam1 (f1); subtract f0 width
                                       from every x-coordinate in pts

    Falls back to det["crop"] (stitched warp) if individual frames are
    unavailable, OBB points are missing, or the re-warp raises an exception.
    """
    pts = det.get("pts")
    if pts is None or f0 is None:
        return det["crop"]

    lx1, _, lx2, _ = det["bbox"]
    cx       = (lx1 + lx2) / 2
    f0_width = f0.shape[1]

    try:
        if cx < f0_width:
            # QR lives in cam0's half — coordinates map directly to f0
            print(f"  [L2] Re-cropping QR from cam0 frame (cx={cx:.0f} < f0_w={f0_width})")
            return get_perspective(f0, pts, pad=L1_PAD_QR, qr_code=True)
        elif f1 is not None:
            # QR lives in cam1's half — shift x-coords into f1's space
            pts_adj = pts.copy().astype("float32")
            pts_adj[:, 0] -= f0_width
            print(f"  [L2] Re-cropping QR from cam1 frame (cx={cx:.0f} ≥ f0_w={f0_width})")
            return get_perspective(f1, pts_adj, pad=L1_PAD_QR, qr_code=True)
    except Exception as exc:
        print(f"  [L2] Individual-frame re-crop failed — falling back to stitched crop: {exc}")

    return det["crop"]


# ── Routes ────────────────────────────────────────────────────────────────────

@app.get("/health")
def health(current_user = Depends(get_current_user)):
    return {"status": "ok", "db_entries": len(medicine_db)}


# ── /medicines ────────────────────────────────────────────────────────────────

@app.get("/medicines")
def get_medicines(current_user = Depends(get_current_user)):
    return {"medicines": expected_list}


@app.post("/medicines")
def set_medicines(payload: ExpectedListPayload, current_user = Depends(get_current_user)):
    global expected_list
    expected_list = [{"name": m.name.upper(), "quantity": m.quantity}
                     for m in payload.medicines]
    return {"medicines": expected_list}


@app.delete("/medicines")
def clear_medicines(current_user = Depends(get_current_user)):
    global expected_list
    expected_list = []
    return {"medicines": []}



# ── /scan ─────────────────────────────────────────────────────────────────────

@app.post("/scan")
def scan(req: ScanRequest, current_user = Depends(get_current_user)):

    t_start = time.time()
    # Individual rotated frames — used by _get_qr_crop for Layer 2.
    # Only populated when two frames arrive separately (frames_b64 path).
    f0_rotated: np.ndarray | None = None
    f1_rotated: np.ndarray | None = None

    try:
        if req.frame_b64:
            # Already stitched by the frontend — no individual frames available;
            # Layer 2 will fall back to using the stitched-frame crop.
            stitched = b64_to_frame(req.frame_b64)

        elif req.frames_b64:
            frames = [b64_to_frame(f) for f in req.frames_b64 if f]
            if len(frames) == 0:
                raise HTTPException(status_code=400, detail="No valid frames provided")
            # stitch_frames returns (stitched, f0, f1) so Layer 2 can
            # re-crop QR detections from the individual camera frames.
            stitched, f0_rotated, f1_rotated = stitch_frames(frames)

        else:
            # No frame from browser — capture directly from the backend cameras
            raw = cam_module.get_raw_frames()
            if raw is None:
                raise HTTPException(status_code=503, detail="Camera not ready — no frames yet")
            f0_raw, f1_raw = raw
            frames = [f for f in [f0_raw, f1_raw] if f is not None]
            stitched, f0_rotated, f1_rotated = stitch_frames(frames)

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid image: {e}")

    t_frames = time.time()
    print(f"[TIMER] Frame decode:     {t_frames - t_start:.2f}s")

    # ── Layer 1 — shape detection on the stitched frame ───────────────────────
    # Single YOLO call at imgsz=3840 so each camera half stays at full
    # 1920×1080 resolution.  All bbox coordinates are already in stitched space
    # so no x-offset adjustment is needed.
    boxes, contents = [], []

    annotated, dets = layer1_detect(stitched, layer1_model, medicine_db)

    for d in dets:
        if d["cls"] == BOX_CLASS_ID:
            boxes.append({
                "box_id": len(boxes) + 1,
                "bbox":   d["bbox"],
                "conf":   d["conf"],
                "qrs":    [],
                "labels": [],
            })
        else:
            contents.append(d)

    t_l1 = time.time()
    print(f"[TIMER] Layer 1:          {t_l1 - t_frames:.2f}s  ({len(boxes)} boxes, {len(contents)} labels/QRs)")

    # ── Layers 2 & 3 — QR and OCR (sequential) ────────────────────────────────
    for d in contents:
        if d["cls"] == QR_CLASS_ID:
            # Re-crop from the individual camera frame for a cleaner QR crop;
            # every other layer still operates on the full stitched frame.
            qr_crop      = _get_qr_crop(d, f0_rotated, f1_rotated)
            data_text    = layer2_read_qr(qr_crop) or "[decode failed]"
            layer3_score = 0.0
        else:
            result = layer3_read_label(d["crop"], medicine_db)
            if result:
                data_text, layer3_score = result
            else:
                data_text    = "[no text found]"
                layer3_score = 0.0

        lx1, ly1, lx2, ly2 = d["bbox"]
        cx, cy = (lx1 + lx2) / 2, (ly1 + ly2) / 2

        best_box  = None
        best_dist = float("inf")
        for b in boxes:
            bx1, by1, bx2, by2 = b["bbox"]
            if bx1 <= cx <= bx2 and by1 <= cy <= by2:
                bcx  = (bx1 + bx2) / 2
                bcy  = (by1 + by2) / 2
                dist = ((cx - bcx) ** 2 + (cy - bcy) ** 2) ** 0.5
                if dist < best_dist:
                    best_dist = dist
                    best_box  = b

        if best_box:
            if d["cls"] == QR_CLASS_ID:
                best_box["qrs"].append(data_text)
            else:
                best_box["labels"].append((data_text, layer3_score))

    t_l23 = time.time()
    print(f"[TIMER] Layer 2/3 OCR:    {t_l23 - t_l1:.2f}s")

    # ── Layer 4 — deferred (only when OCR is weak) ────────────────────────────
    layer4_detections = None

    def get_layer4():
        nonlocal layer4_detections
        if layer4_detections is None:

            individual = [f for f in [f0_rotated, f1_rotated] if f is not None]
            layer4_detections = layer4_scan_full_frame(
                stitched,
                original_frames=individual if len(individual) > 1 else None,
            )
        return layer4_detections

    expected_qty: dict = {}
    for item in req.expected:
        if isinstance(item, str):
            n, q = item.upper().strip(), 1
        else:
            n = item.get("name", "").upper().strip()
            q = int(item.get("quantity", 1))
        if n:
            expected_qty[n] = expected_qty.get(n, 0) + q


    found_counts: dict = {}
    results = []

    for b in boxes:
        # Labels stored as (texts, layer3_db_score) tuples.
        # Unpack texts for best_ocr_label, track highest layer3 score per box.
        all_ocr       = []
        best_l3_score = 0.0
        for lbl in b["labels"]:
            if isinstance(lbl, tuple):
                texts, l3_score = lbl
                try:
                    best_l3_score = max(best_l3_score, float(np.array(l3_score).flat[0]))
                except Exception:
                    pass
                items = texts if isinstance(texts, list) else [texts]
            elif isinstance(lbl, list):
                items = lbl
            else:
                items = [lbl]
            all_ocr.extend(items)

        all_qr = [qr for qr in b["qrs"] if qr and qr != "[decode failed]"]
        ocr_in = best_ocr_label(all_ocr)

        ocr_db_name, ocr_db_score = (None, 0.0)
        if ocr_in:
            ocr_db_name, ocr_db_score = db_match(
                normalize_ocr(ocr_in.lower()), medicine_db)

        # If layer3 internal score is higher than re-scoring raw text, use it.
        if best_l3_score > ocr_db_score:
            print(f"  [L3 score boost] layer3={best_l3_score:.2f} > rescore={ocr_db_score:.2f}")
            ocr_db_score = best_l3_score

        qr_present = len(all_qr) > 0
        ocr_high   = db_confidence_tier(ocr_db_score) == "HIGH"

        if not qr_present and not ocr_high:
            dets = get_layer4()
            vision_name, vision_conf = layer4_match_to_box(dets, b["bbox"])
        else:
            vision_name, vision_conf = "UNKNOWN", 0.0

        verdict = consensus_check(
            ocr_texts    = [ocr_in],
            qr_texts     = all_qr,
            vision_name  = vision_name,
            vision_conf  = vision_conf,
            medicine_db  = medicine_db,
            ocr_db_name  = ocr_db_name,
            ocr_db_score = ocr_db_score,
        )

        name = verdict["final_name"].upper()

        if verdict["final_name"] == "PENDING_REVIEW":
            scan_status = "PENDING_REVIEW"
        elif verdict["final_name"] == "UNKNOWN":
            scan_status = "UNKNOWN"
        elif name in expected_qty:
            found_counts[name] = found_counts.get(name, 0) + 1
            if found_counts[name] <= expected_qty[name]:
                scan_status = "MATCHED"     # within expected quantity
            else:
                scan_status = "EXTRA"       # found more than expected
        else:
            scan_status = "EXTRA"           # not in expected list

        results.append({
            "box_id":          b["box_id"],
            "bbox":            list(b["bbox"]),
            "final_name":      verdict["final_name"],
            "confidence":      verdict["confidence"],
            "layer4_note":     verdict["layer4_note"],
            "status":          verdict["status"],
            "scan_status":     scan_status,
            "ocr_raw":         verdict["ocr_raw"],
            "qr_name":         verdict["qr_name"],
            "reference_image": None,
        })

    # Draw UNKNOWN boxes on the annotated frame so the pharmacist can see
    # which physical box the system couldn't identify
    for r in results:
        if r["scan_status"] != "UNKNOWN" or not r["bbox"]:
            continue
        x1, y1, x2, y2 = r["bbox"]

        # Bright orange border — distinct from YOLO's default colours
        cv2.rectangle(annotated, (x1, y1), (x2, y2), (0, 140, 255), 3)

        # "? UNKNOWN" label with filled background for readability
        label      = "? UNKNOWN"
        font       = cv2.FONT_HERSHEY_SIMPLEX
        font_scale = 0.8
        thickness  = 2
        (tw, th), baseline = cv2.getTextSize(label, font, font_scale, thickness)

        # Clamp label above box — don't let it go off screen
        label_y = max(y1, th + baseline + 6)
        cv2.rectangle(annotated,
                      (x1, label_y - th - baseline - 6),
                      (x1 + tw + 6, label_y),
                      (0, 140, 255), -1)
        cv2.putText(annotated, label,
                    (x1 + 3, label_y - baseline - 3),
                    font, font_scale, (255, 255, 255), thickness)

    # Mark MISSING — one entry PER missing unit so the list count
    # matches the summary number (3 missing Sefloc = 3 list entries).
    for name, qty_needed in expected_qty.items():
        qty_found   = found_counts.get(name, 0)
        qty_missing = qty_needed - qty_found
        for i in range(qty_missing):
            results.append({
                "box_id":          None,
                "bbox":            None,
                "final_name":      name,
                "confidence":      "NONE",
                "layer4_note":     f"Expected {qty_needed}× — only {qty_found} found",
                "status":          "❌ Not detected",
                "scan_status":     "MISSING",
                "ocr_raw":         "",
                "qr_name":         None,
                "reference_image": find_medicine_image(name),
                "qty_expected":    qty_needed,
                "qty_found":       qty_found,
                "qty_missing":     qty_missing,
                "unit_index":      i + 1,
            })

    # Summary counts
    matched = sum(1 for r in results if r["scan_status"] == "MATCHED")
    missing = sum(1 for r in results if r["scan_status"] == "MISSING")
    extra   = sum(1 for r in results if r["scan_status"] == "EXTRA")
    review  = sum(1 for r in results if r["scan_status"] == "PENDING_REVIEW")
    unknown = sum(1 for r in results if r["scan_status"] == "UNKNOWN")

    if missing > 0:
        led_red()
        threading.Thread(target=blink_then_white, args=(led_red,), daemon=True).start()
    elif extra > 0:
        led_orange()
        threading.Thread(target=blink_then_white, args=(led_orange,), daemon=True).start()
    elif unknown > 0:
        led_unknown()
        threading.Thread(target=blink_then_white, args=(led_unknown,), daemon=True).start()
    else:
        led_green()
        threading.Thread(target=blink_then_white, args=(led_green,), daemon=True).start()

    
    ts       = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = os.path.join(LOG_DIR, f"scan_result_{ts}.jpg")
    cv2.imwrite(out_path, annotated)

    
    l4_out_path      = None
    l4_annotated_b64 = None
    if layer4_detections is not None:
        from pipeline.layer4_vision import layer4_draw_annotated

        l4_img           = layer4_draw_annotated(stitched, layer4_detections)
        l4_out_path      = os.path.join(LOG_DIR, f"scan_layer4_{ts}.jpg")
        cv2.imwrite(l4_out_path, l4_img)
        l4_annotated_b64 = frame_to_b64(l4_img)
        print(f"[Scan] Saved L4 debug (stitched) → {l4_out_path}")

    # Add to scan history
    history_entry = {
        "id":        str(uuid.uuid4()),
        "timestamp": datetime.now().isoformat(),
        "matched":   matched,
        "missing":   missing,
        "extra":     extra,
        "review":    review,
        "medicines": [r["final_name"] for r in results
                      if r["scan_status"] == "MATCHED"],
        "summary":   results[0]["final_name"] if results else "No medicines",
    }
    scan_history.insert(0, history_entry)
    if len(scan_history) > 50:
        scan_history.pop()

    t_end = time.time()
    print(f"[TIMER] Layer 4+consensus:{t_end - t_l23:.2f}s")
    print(f"[TIMER] -- TOTAL -------- {t_end - t_start:.2f}s")

    return {
        "results":        results,
        "summary":        {"matched": matched, "missing": missing,
                           "extra": extra, "review": review},
        "annotated_b64":    frame_to_b64(annotated),
        "l4_annotated_b64": l4_annotated_b64,
        "saved_to":         out_path,
        "timestamp":      datetime.now().isoformat(),
    }


@app.get("/scan/history")
def get_history(current_user = Depends(get_current_user)):
    return {"history": scan_history}


@app.delete("/scan/history")
def clear_history(current_user = Depends(get_current_user)):
    scan_history.clear()
    return {"history": []}